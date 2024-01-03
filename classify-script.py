import os
import openpyxl

def classify_luat_dau_thau(sheet):
    # Get the last row with data
    last_row = sheet.max_row

    # Find the column index based on the column name

    TTBYT_keywords = ["thiết bị", "máy", "vật tư", "dụng cụ", "bơm", "bông", "băng", "gạc", "kim", "dây", "túi", "Catheter", "chỉ", "dao", "van", "giá đỡ", "stent", "đầu", "xương", "sụn", "khớp", "gân", "thuỷ tinh", "miếng", "mảnh", "màng", "bao", "que", "ống", "vật liệu", "nong", "cỡ", "bộ", "phim", "bóng", "mũi", "máy", "quả"]
    vattu_xn_keywords = ["Xét nghiệm", "bộ test", "đĩa", "khay", "kit", "kít", "test"]
    hoachat_keywords = ["hoá chất", "chất thử", "chất chuẩn", "chất kháng", "chất hiệu chuẩn", "chất kiểm", "dung dịch", "anti", "định lượng", "đối chiếu", "đối chứng", "thuốc thử", "môi trường", "bột", "chất", "sinh phẩm", "acid", "axit", "khí", "dịch", "dd (viết tắt của dung dịch)", "kháng thể", "kháng nguyên"]
    thuoc_keywords = ["Thuốc", "hoạt chất", "Vắc xin"]
    khac_keywords = ["Cung cấp", "lắp đặt", "bảo trì", "bảo dưỡng", "sửa chữa"]

    for row_number in range(3, last_row + 1):
        cell_value_l = sheet[f"L{row_number}"].value

        # Check if column L contains any of the specified keywords
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in TTBYT_keywords):
            # Update the corresponding row in column AD to "TTBYT"
            sheet[f"AD{row_number}"] = "TTBYT"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in vattu_xn_keywords):
            # Update the corresponding row in column AD to "Vật tư XN"
            sheet[f"AD{row_number}"] = "Vật tư XN"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in hoachat_keywords):
            # Update the corresponding row in column AD to "Hoá chất"
            sheet[f"AD{row_number}"] = "Hoá chất"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khac_keywords):
            # Update the corresponding row in column AD to "Khác"
            sheet[f"AD{row_number}"] = "Khác"

def classify_chuyenkhoa(sheet):
    # Get the last row with data
    last_row = sheet.max_row

    khoa_than_keywords = ["Thận", "lọc máu", "dịch lọc", "thẩm phân", "FAV", "quả lọc", "màng lọc", "peracetic", "hd plus", "axit citric", "acid citric", "lọc màng bụng", "phúc mạc", "TNT"]
    khoa_rang_ham__mat_keywords = ["Nha khoa", "răng", "ống tuỷ", "nướu", "Composite", "Cone", "trám", "đánh bóng", "nước bọt", "trâm", "mặt", "hàm"]
    khoa_tieu_hoa_keywords = ["tiêu hoá", "dạ dày", "trực tràng", "thực quản", "hậu môn", "miệng", "đại tràng", "polyp", "đường mật"]
    khoa_xet_nghiem_keywords = ["Xét nghiệm", " khoanh", " đĩa", " môi trường", " khánh sinh", " kháng thể", " thuốc thử", " chất thử", " thử nghiệm", " định danh", " nhuộm", " hồng cầu", " định tính", " định lượng", " lam kính", " bệnh phẩm", " ống nghiệm", " test", " que thử", " lame", " phản ứng", " pha loãng", " huyết học", " tế bào", " hiệu chuẩn", " giếng"]
    nhan_khoa_keywords = ["Nhãn khoa", "mắt", "thuỷ tinh thể"]
    khoa_chan_doan_hinh_anh_keywords = ["Phim", "X-quang", "X-ray", "film", "siêu âm", "điện tim"]
    khoa_xuong_khop_keywords = ["Băng bột bó", "xương", "khớp", "chỉnh hình", "đai", "nẹp", "chân", "tay", "đùi", "đinh", "vít", "khung", "gân", "đốt sống", "đĩa đệm", "cột sống", "garo", "băng bó bột", "cổ", "sụn"]
    san_khoa_keywords = ["màng cứng", "thai", "sản", "âm đạo", "sơ sinh"]
    nhi_khoa_keywords = ["trẻ em", "nhi khoa", "khoa nhi"]
    gay_me_keywords = ["Gây mê", "gây tê", "mask", "mặt nạ"]
    khoa_tim_keywords = ["Tim"]
    tai_mui_hong_keywords = ["Tai", "mũi", "họng", "khí quản", "lưỡi", "đàm"]
    ngoai_khoa_keywords = ["Ngoại", "mổ", "phẫu thuật", "cầm máu", "chỉ", "dao", "cắt", "khâu", "nội soi", "kẹp", "kiềm", "dị vật"]
    khoa_tiet_nieu_keywords = ["niệu quản", "tiểu", "ống thông"]
    khoa_ho_hap_keywords = ["Phổi", "khí quản", "ngực", "thở"]
    khoa_than_kinh_keywords = ["Sọ", "Não"]
    chung_keywords = ["Sát khuẩn", "khử khuẩn", "cồn", "nước cất", "bông", "băng", "gạc", "bơm tiêm", "kim tiêm", "dây truyền máu", "dây truyền dịch", "dây nối", "khoá", "găng tay", " găng", "túi ép", "túi máu", "khẩu trang", "rửa", "tẩy", "làm sạch", "tiệt khuẩn", "diệt khuẩn", "khử trùng", "dây cho ăn"]
    for row_number in range(3, last_row + 1):
        cell_value_l = sheet[f"L{row_number}"].value
        # Check if column L contains any of the specified keywords
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_than_keywords):
            # Update the corresponding row in column AE to "Khoa Thận"
            sheet[f"AE{row_number}"] = "Khoa Thận"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_rang_ham__mat_keywords):
            # Update the corresponding row in column AE to "Khoa Răng hàm mặt"
            sheet[f"AE{row_number}"] = "Khoa Răng hàm mặt"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_tieu_hoa_keywords):
            # Update the corresponding row in column AE to "Khoa Tiêu hoá"
            sheet[f"AE{row_number}"] = "Khoa Tiêu hoá"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_xet_nghiem_keywords):
            # Update the corresponding row in column AE to "Khoa Xét nghiệm"
            sheet[f"AE{row_number}"] = "Khoa Xét nghiệm"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in nhan_khoa_keywords):
            # Update the corresponding row in column AE to "Nhãn khoa"
            sheet[f"AE{row_number}"] = "Nhãn khoa"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_chan_doan_hinh_anh_keywords):
            # Update the corresponding row in column AE to "Khoa Chuẩn đoán hình ảnh"
            sheet[f"AE{row_number}"] = "Khoa Chuẩn đoán hình ảnh"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_xuong_khop_keywords):
            # Update the corresponding row in column AE to "Khoa Xương khớp"
            sheet[f"AE{row_number}"] = "Khoa Xương khớp"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in san_khoa_keywords):
            # Update the corresponding row in column AE to "Sản khoa"
            sheet[f"AE{row_number}"] = "Sản khoa"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in nhi_khoa_keywords):
            # Update the corresponding row in column AE to "Nhi khoa"
            sheet[f"AE{row_number}"] = "Nhi khoa"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in gay_me_keywords):
            # Update the corresponding row in column AE to "Khoa Gây mê, Hồi sức"
            sheet[f"AE{row_number}"] = "Khoa Gây mê, Hồi sức"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_tim_keywords):
            # Update the corresponding row in column AE to "Khoa Tim mạch"
            sheet[f"AE{row_number}"] = "Khoa Tim mạch"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in tai_mui_hong_keywords):
            # Update the corresponding row in column AE to "Khoa Tai mũi họng"
            sheet[f"AE{row_number}"] = "Khoa Tai mũi họng"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in ngoai_khoa_keywords):
            # Update the corresponding row in column AE to "Ngoại khoa"
            sheet[f"AE{row_number}"] = "Ngoại khoa"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_tiet_nieu_keywords):
            # Update the corresponding row in column AE to "Khoa tiết niệu"
            sheet[f"AE{row_number}"] = "Khoa tiết niệu"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_ho_hap_keywords):
            # Update the corresponding row in column AE to "Khoa Hô hấp"
            sheet[f"AE{row_number}"] = "Khoa Hô hấp"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in khoa_than_kinh_keywords):
            # Update the corresponding row in column AE to "Khoa Thần kinh"
            sheet[f"AE{row_number}"] = "Khoa Thần kinh"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in chung_keywords):
            # Update the corresponding row in column AE to "Chung"
            sheet[f"AE{row_number}"] = "Chung"

def classify_dongsp_chuyenkhoa(sheet):
    # Get the last row with data
    last_row = sheet.max_row

    qua_loc_lientuc_pair_keywords = [("quả", "liên tục"), ("màng", "liên tục"), ("Quả", "omni"), ("màng", "omni"), ("Quả", "bộ dây"), ("màng", "Bộ dây"), ("quả", "prismaflex"), ("màng", "prismaflex")]
    qua_loc_happhu_pair_keywords = [("quả", "hấp phụ"), ("màng", "hấp phụ")]
    qua_loc_tachhuyettuong_pair_keywords = [("quả", "huyết tương")]
    qua_loc_chuky_keywords = ["Quả lọc", "màng lọc", "siêu lọc", "thông lượng", "low", "high", "middle"]
    loc_mang_bung_keywords = ["Màng bụng"]
    day_loc_lientuc_pair_keywords = [("Dây", "liên tục"), ("dây", "prismaflex")]
    day_keywords = ["Dây"]
    kim_keywords = ["Kim"]
    dich_loc_keywords = ["Dịch lọc", "thẩm phân", "HD plus"]
    axit_citric_bot_keywords = ["Axit citric", " acid citric", " rửa máy", " tẩy rửa máy", " khử trùng máy", " khử khuẩn máy", " làm sạch máy", " tẩy khuẩn máy"]
    axit_citric_long_keywords = ["Axit citric", " acid citric", " rửa máy", " tẩy rửa máy", " khử trùng máy", " khử khuẩn máy", " làm sạch máy", " tẩy khuẩn máy"]
    dd_rua_mang_keywords = ["Rửa màng", "rửa quả", "peracetic", "MDT", "vertecid", "vertexit", "khử trùng quả", "khử trùng màng", "bảo quản quả", "bảo quản màng", "diệt khuẩn màng", "diệt khuẩn quả", "ngâm quả", "ngâm màng", "sát khuẩn màng", "sát khuẩn quả", "tẩy khuẩn màng", "tẩy khuẩn quả", "tẩy trùng quả", "tẩy trùng màng"]
    may_hdf_online_pair_keywords = [("Máy", "HDF")]
    may_than_lien_tuc_pair_keywords = [("Máy", "liên tục"), ("máy", "prismaflex")]
    may_than_chu_ky_keywords = ["máy chạy thận", "máy lọc thận", "máy thận", "máy TNT"]
    may_ro_pair_keywords = [("Máy", "RO")]
    may_rua_pair_keywords = [("Máy", "rửa")]
    fav_keywords = ["FAV", "bộ tiêm chích", "gạc thận"]

    for row_number in range(3, last_row + 1):
        cell_value_l = sheet[f"L{row_number}"].value
        sheet[f"AF{row_number}"] = "Khác"
        # Check if column L contains any of the specified keywords
        for pair in qua_loc_lientuc_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Quả lọc liên tục"
                sheet[f"AF{row_number}"] = "Quả lọc liên tục"
        for pair in qua_loc_happhu_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Quả lọc hấp phụ"
                sheet[f"AF{row_number}"] = "Quả lọc hấp phụ"
        for pair in qua_loc_tachhuyettuong_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Quả lọc tách huyết tương"
                sheet[f"AF{row_number}"] = "Quả lọc tách huyết tương"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in qua_loc_chuky_keywords):
            # Update the corresponding row in column AF to "Quả lọc chu kỳ"
            sheet[f"AF{row_number}"] = "Quả lọc chu kỳ"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in loc_mang_bung_keywords):
            # Update the corresponding row in column AF to "Lọc màng bụng"
            sheet[f"AF{row_number}"] = "Lọc màng bụng"
        for pair in day_loc_lientuc_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Dây lọc liên tục"
                sheet[f"AF{row_number}"] = "Dây lọc liên tục"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in day_keywords):
            # Update the corresponding row in column AF to "Dây lọc chu kỳ"
            sheet[f"AF{row_number}"] = "Dây lọc chu kỳ"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in kim_keywords):
            # Update the corresponding row in column AF to "Kim"
            sheet[f"AF{row_number}"] = "Kim"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in dich_loc_keywords):
            # Update the corresponding row in column AF to "Dịch lọc"
            sheet[f"AF{row_number}"] = "Dịch lọc"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in axit_citric_bot_keywords):
            # Update the corresponding row in column AF to "Axit citric bột"
            sheet[f"AF{row_number}"] = "Axit citric bột"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in axit_citric_long_keywords):
            # Update the corresponding row in column AF to "Axit citric lỏng"
            sheet[f"AF{row_number}"] = "Axit citric lỏng"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in dd_rua_mang_keywords):
            # Update the corresponding row in column AF to "Dung dịch rửa màng"
            sheet[f"AF{row_number}"] = "Dung dịch rửa màng"
        for pair in may_hdf_online_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Máy HDF online"
                sheet[f"AF{row_number}"] = "Máy HDF online"
        for pair in may_than_lien_tuc_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Máy thận liên tục"
                sheet[f"AF{row_number}"] = "Máy thận liên tục"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in may_than_chu_ky_keywords):
            # Update the corresponding row in column AF to "Máy thận chu kỳ"
            sheet[f"AF{row_number}"] = "Máy thận chu kỳ"
        for pair in may_ro_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Máy RO"
                sheet[f"AF{row_number}"] = "Máy RO"
        for pair in may_rua_pair_keywords:
            if pair[0]in cell_value_l and pair[1]in cell_value_l:
                # Update the corresponding row in column AF to "Máy rửa"
                sheet[f"AF{row_number}"] = "Máy rửa"
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in fav_keywords):
            # Update the corresponding row in column AF to "FAV"
            sheet[f"AF{row_number}"] = "FAV"


def classify_5086(sheet):
    last_row = sheet.max_row
    xquang_keywords = ["Phim X Quang", "Phim X-Quang", "Phim XQ", "Phim Xquang", "Phim khô", "phim xquang", "phim xray", "phim x-ray", "Film khô", "Phim Cea- di răng", "Phim Cea- di răng", "Phim chụp răng", "Phim lọc", "Phim chụp laser", "Film", "Phim rửa liền X ray"]
    for row_number in range(3, last_row + 1):
        cell_value_l = sheet[f"L{row_number}"].value
        if isinstance(cell_value_l, str) and any(keyword.lower() in cell_value_l.lower() for keyword in xquang_keywords):
            # Update the corresponding row in column AG to "Tim mạch và X- quang can thiệp"
            sheet[f"AG{row_number}"] = "Tim mạch và X- quang can thiệp"

def classify(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # Phan loai theo luat dau thau
    classify_luat_dau_thau(sheet)
    print("Phan loai theo luat dau thau xong!")

    # Phan loai theo chuyen khoa
    classify_chuyenkhoa(sheet)
    print("Phan loai theo chuyen khoa xong!")

    # Phan loai theo sp chuyen khoa
    classify_dongsp_chuyenkhoa(sheet)
    print("Phan loai theo dong sp chuyen khoa xong!")

    # Phan loai theo 5086
    classify_5086(sheet)
    print("Phan loai theo 5086 xong!")

    print("Script completed.")
    
    workbook.save(file_path)
    workbook.close()

if __name__ == "__main__":
    #excel_file_path = "C:\\Users\\nnt12\\Documents\\Excel\\sample.xlsx"
    #excel_file_path = "D:\\Document\\Excel\\classify_sample.xlsx"
    print("Start to classify.")
    # Get the current working directory
    folder_path = os.getcwd()

    # Get the list of files in the folder
    files = os.listdir(folder_path)

    # Filter out the Excel files
    excel_files = [file for file in files if file.endswith('.xlsx') or file.endswith('.xls')]

    # Open each Excel file
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        sheet_name = "sheet"
        classify(file_path, sheet_name)
        print("Classsify file done.")
    print("Classsify all files done.")
    
    # Wait for user input before closing the terminal
    input("Press Enter to close the terminal.")